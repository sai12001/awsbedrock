import boto3
import openpyxl
import csv
import io
from io import BytesIO
import base64
import json
import logging

logging.basicConfig(level=logging.DEBUG)

def process_event(prompt, file_contents, filetype):
    try:
        logging.debug(f"Prompt: {prompt}")
        logging.debug(f"Filetype: {filetype}")

        if file_contents:
            logging.debug("Processing file contents...")
            if filetype.lower() in ['xlsx', 'xls',"vnd.openxmlformats-officedocument.spreadsheetml.sheet"]:
                # Read Excel file
                excel_file = BytesIO(file_contents)
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active

                # Process Excel data
                excel_info = []
                for row in sheet.iter_rows(values_only=True):
                    excel_info.append(list(row))

                # Convert Excel data to a readable string format
                excel_data_text = "\n".join([", ".join(map(str, row)) for row in excel_info])
                logging.debug(f"Excel data text: {excel_data_text}")

                # Add Excel data as text to the request body
                prompt += f"\ndata frame:\n{excel_data_text}"

            elif filetype.lower() == 'csv':
                # Read CSV file
                csv_data = file_contents.decode('utf-8')
                csv_reader = csv.reader(io.StringIO(csv_data))

                # Process CSV rows
                csv_info = []
                for row in csv_reader:
                    csv_info.append(row)

                # Convert CSV data to a readable string format
                csv_data_text = "\n".join([", ".join(map(str, row)) for row in csv_info])
                logging.debug(f"CSV data text: {csv_data_text}")

                # Add CSV data as text to the request body
                prompt += f"\nData Frame:\n{csv_data_text}"
            
            else:
                logging.debug(f"Unhandled file type: {filetype}")

        else:
            logging.debug("No file contents provided, using S3 file.")
            # Simulate fetching a file from S3 if no file is uploaded
            s3_client = boto3.client('s3')
            bucket_name = 'bedrocktest03'
            file_key = 'Employee_Details-2.xlsx'  # Adjust file extension based on 'filetype'
            s3_object = s3_client.get_object(Bucket=bucket_name, Key=file_key)
            file_data = s3_object['Body'].read()

            # Assume the file is an Excel file for this example
            excel_file = BytesIO(file_data)
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            excel_info = []
            for row in sheet.iter_rows(values_only=True):
                excel_info.append(list(row))
            excel_data_text = "\n".join([", ".join(map(str, row)) for row in excel_info])
            logging.debug(f"Excel data text from S3: {excel_data_text}")
            prompt += f"\nExcel file contents:\n{excel_data_text}"
        
        prompt += "Gets the information from the given Query from the Dataframe, if the query is realted to manipulation and the answer should be in 3 lines don't provide any code"
        # Create a request body for Bedrock
        request_body = {
            "anthropic_version": "bedrock-2023-05-31",
            "max_tokens": 900,
            "messages": [
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "text",
                            "text": prompt
                        }
                    ]
                }
            ]
        }

        logging.debug(f"Request Body: {json.dumps(request_body)}")

        # Create a Bedrock runtime client (replace with your region)
        client = boto3.client('bedrock-runtime', region_name='us-east-1')

        # Set the model ID for Claude 3 Sonnet
        model_id = "anthropic.claude-3-sonnet-20240229-v1:0"

        # Send the request to Bedrock using the 'invoke' API method
        response = client.invoke_model(
            modelId=model_id,
            body=json.dumps(request_body)
        )

        # Read the content from the StreamingBody
        response_payload = response['body'].read().decode('utf-8')
        logging.debug(f"Response Payload: {response_payload}")

        payload = json.loads(response_payload)
        generated_text = payload.get('completions', [{}])[0].get('text', '')

        # Return the generated text
        return {
            'generated_text': generated_text,
            'response': payload
        }

    except Exception as e:
        logging.error(f"Error: {str(e)}")
        return {
            'error': str(e)
        }
